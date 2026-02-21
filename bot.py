"""Telegram bot for ARGSMS - SMS range management system."""

import os
import json
import logging
import random
import re
import traceback
from datetime import datetime, timedelta
from io import BytesIO
from dotenv import load_dotenv

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    MessageHandler,
    filters,
    ConversationHandler
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from database import (
    init_db,
    get_or_create_user,
    log_access,
    is_user_admin,
    is_user_banned,
    get_user_balance,
    add_user_balance,
    deduct_user_balance,
    get_price_for_range,
    create_number_holds,
    mark_number_permanent,
    get_held_numbers,
    is_number_held,
    cleanup_expired_holds,
    update_first_retry_time,
    User,
    NumberHold,
    PriceRange,
    Transaction,
    RechargeRequest,
    AccessLog,
    Range,
    PhoneNumber,
    import_csv_data,
    get_all_ranges,
    get_range_by_unique_id,
    get_available_numbers_for_range,
    delete_range_and_numbers,
    set_range_price
)
from scrapper_wrapper import get_scrapper_session

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Get bot token from environment
BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
if not BOT_TOKEN:
    raise ValueError("TELEGRAM_BOT_TOKEN must be set in .env file")

# Get admin username for recharge contact
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "")

# Get force join channel ID
FORCE_JOIN_CHANNEL_ID = os.getenv("FORCE_JOIN_CHANNEL_ID", "")

# Constants for button text formatting
MAX_BUTTON_TEXT_LENGTH = 60
MAX_TITLE_LENGTH = 50

# Constants for SMS number fetching
SMS_FETCH_COUNT = 100  # How many numbers to fetch from API
SMS_DISPLAY_COUNT = 20  # How many random numbers to display

# Constants for SMS message display
MAX_TELEGRAM_MESSAGE_LENGTH = 3500  # Leave room for additional text

# Conversation states for admin operations
WAITING_FOR_ADD_BALANCE_AMOUNT = 1
WAITING_FOR_DEDUCT_BALANCE_AMOUNT = 2
WAITING_FOR_PRICE_PATTERN = 3
WAITING_FOR_PRICE_AMOUNT = 4
WAITING_FOR_CSV_FILE = 5

# Initialize database session factory
SessionFactory = init_db()


def get_db_session():
    """Get a new database session."""
    return SessionFactory()


def escape_html(text):
    """Escape HTML special characters for safe display in Telegram messages."""
    if text is None:
        return ""
    return str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def strip_html_tags(html_str):
    """
    Strip HTML tags from string and decode HTML entities.
    
    Args:
        html_str: String potentially containing HTML tags
    
    Returns:
        Clean text without HTML tags
    """
    if not html_str or not isinstance(html_str, str):
        return str(html_str) if html_str else ""
    
    # Use a simple regex to remove HTML tags
    import re
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', html_str)
    # Decode common HTML entities
    text = text.replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&')
    text = text.replace('&quot;', '"').replace('&#039;', "'").replace('&nbsp;', ' ')
    # Clean up whitespace
    text = ' '.join(text.split())
    return text.strip()


async def check_channel_membership(update: Update, context: ContextTypes.DEFAULT_TYPE, db_session=None):
    """
    Check if user is a member of the required channel.
    Returns True if user is a member or no channel is configured, False otherwise.
    If False, sends a message to the user with a join button.
    Admins bypass this check.
    """
    # If no channel is configured, skip check
    if not FORCE_JOIN_CHANNEL_ID:
        return True
    
    user_id = update.effective_user.id
    
    # Check if user is admin - admins bypass channel check
    if db_session and is_user_admin(db_session, user_id):
        return True
    
    try:
        # Check if user is a member of the channel
        member = await context.bot.get_chat_member(FORCE_JOIN_CHANNEL_ID, user_id)
        
        # Check if user has left or is kicked
        if member.status in ['left', 'kicked']:
            # User is not a member, send join prompt
            channel_name = FORCE_JOIN_CHANNEL_ID[1:] if FORCE_JOIN_CHANNEL_ID.startswith('@') else FORCE_JOIN_CHANNEL_ID
            keyboard = [[InlineKeyboardButton("📢 Join Channel", url=f"https://t.me/{channel_name}")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            message_text = (
                "🔒 <b>Channel Membership Required</b>\n\n"
                "To use this bot, you must join our channel first.\n"
                "Click the button below to join, then try again."
            )
            
            # Send or edit message depending on context
            if update.callback_query:
                await update.callback_query.answer("❌ Please join our channel first", show_alert=True)
                await update.callback_query.edit_message_text(
                    message_text,
                    reply_markup=reply_markup,
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(
                    message_text,
                    reply_markup=reply_markup,
                    parse_mode='HTML'
                )
            return False
        
        # User is a member (status: creator, administrator, member, restricted)
        return True
        
    except Exception as e:
        logger.error(f"Error checking channel membership: {e}")
        # On error, allow access (fail open)
        return True


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the /start command."""
    user = update.effective_user
    db = get_db_session()
    
    try:
        # Get or create user in database
        db_user = get_or_create_user(db, user.id, user.username)
        
        # Check if user is banned
        if db_user.is_banned:
            await update.message.reply_text(
                "🚫 You have been banned from using this bot.\n"
                "Please contact an administrator if you believe this is an error."
            )
            return
        
        # Check channel membership
        if not await check_channel_membership(update, context, db):
            return
        
        log_access(db, db_user, "start_command")
        
        # Create main menu
        keyboard = [
            [InlineKeyboardButton("📱 View SMS Ranges", callback_data="view_sms_ranges")],
            [InlineKeyboardButton("👤 My Profile", callback_data="user_profile")],
            [InlineKeyboardButton("💰 Recharge Balance", callback_data="recharge_request")],
            [InlineKeyboardButton("ℹ️ About", callback_data="about")]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        welcome_message = (
            f"👋 Welcome to ARGSMS Bot, {user.first_name}!\n\n"
            f"💰 Balance: ${db_user.balance:.2f}\n\n"
            "This bot allows you to view available SMS ranges.\n"
            "Use the menu below to navigate:"
        )
        
        await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    finally:
        db.close()


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle the /admin command for admin panel."""
    user = update.effective_user
    db = get_db_session()
    
    try:
        # Check if user is admin
        if not is_user_admin(db, user.id):
            await update.message.reply_text("❌ You don't have admin privileges.")
            return
        
        db_user = get_or_create_user(db, user.id, user.username)
        log_access(db, db_user, "admin_panel_access")
        
        # Create admin panel keyboard
        keyboard = [
            [InlineKeyboardButton("👥 List Users", callback_data="admin_list_users")],
            [InlineKeyboardButton("🔑 Manage Admins", callback_data="admin_manage_admins")],
            [InlineKeyboardButton("🚫 Ban/Unban Users", callback_data="admin_manage_bans")],
            [InlineKeyboardButton("💰 Manage Balance", callback_data="admin_manage_balance")],
            [InlineKeyboardButton("💳 Recharge Requests", callback_data="admin_recharge_requests")],
            [InlineKeyboardButton("📤 Upload CSV (Ranges & Numbers)", callback_data="admin_upload_csv")],
            [InlineKeyboardButton("📋 Manage Ranges & Prices", callback_data="admin_manage_ranges")],
            [InlineKeyboardButton("🔒 Number Holds Report", callback_data="admin_number_holds")],
            [InlineKeyboardButton("📊 View Stats", callback_data="admin_view_stats")],
            [InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_to_main")]
        ]
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        admin_message = (
            "🔐 Admin Panel\n\n"
            "Welcome to the admin panel. Select an option:"
        )
        
        await update.message.reply_text(admin_message, reply_markup=reply_markup)
    finally:
        db.close()


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle button callbacks from inline keyboards."""
    query = update.callback_query
    user = query.from_user
    
    # Answer the callback query to remove loading state
    await query.answer()
    
    db = get_db_session()
    try:
        # Get or create user
        db_user = get_or_create_user(db, user.id, user.username)
        
        # Check if user is banned (except for back_to_main)
        callback_data = query.data
        if callback_data != "back_to_main" and db_user.is_banned:
            await query.answer("🚫 You have been banned from using this bot.", show_alert=True)
            return
        
        # Check channel membership for non-admin actions
        if not callback_data.startswith("admin_") and callback_data != "back_to_main":
            if not await check_channel_membership(update, context, db):
                return
        
        # Main menu callbacks
        if callback_data == "view_sms_ranges":
            await view_sms_ranges_callback(query, context, db, db_user)
        elif callback_data == "user_profile":
            await user_profile_callback(query, context, db, db_user)
        elif callback_data == "recharge_request":
            await recharge_request_callback(query, context, db, db_user)
        elif callback_data == "about":
            await about_callback(query, context, db, db_user)
        elif callback_data == "back_to_main":
            await back_to_main_callback(query, context, db, db_user)
        
        # Admin panel callbacks
        elif callback_data == "admin_list_users":
            await admin_list_users_callback(query, context, db, db_user)
        elif callback_data == "admin_manage_admins":
            await admin_manage_admins_callback(query, context, db, db_user)
        elif callback_data == "admin_manage_bans":
            await admin_manage_bans_callback(query, context, db, db_user)
        elif callback_data == "admin_manage_balance":
            await admin_manage_balance_callback(query, context, db, db_user)
        elif callback_data == "admin_recharge_requests":
            await admin_recharge_requests_callback(query, context, db, db_user)
        elif callback_data == "admin_upload_csv":
            await admin_upload_csv_callback(query, context, db, db_user)
        elif callback_data == "admin_manage_ranges":
            await admin_manage_ranges_callback(query, context, db, db_user)
        elif callback_data == "admin_view_stats":
            await admin_view_stats_callback(query, context, db, db_user)
        elif callback_data == "admin_number_holds":
            await admin_number_holds_callback(query, context, db, db_user)
        elif callback_data == "admin_export_holds":
            await admin_export_holds_callback(query, context, db, db_user)
        elif callback_data == "admin_cleanup_holds":
            await admin_cleanup_holds_callback(query, context, db, db_user)
        elif callback_data == "admin_release_all_holds":
            await admin_release_all_holds_callback(query, context, db, db_user)
        elif callback_data == "admin_back":
            await admin_back_callback(query, context, db, db_user)
        
        # Range detail view (updated to use range_ prefix instead of sms_range_)
        elif callback_data.startswith("range_"):
            range_unique_id = callback_data.replace("range_", "")
            await view_sms_range_detail_callback(query, context, db, db_user, range_unique_id)
        
        # Set price for a range
        elif callback_data.startswith("set_price_"):
            range_unique_id = callback_data.replace("set_price_", "")
            await admin_set_range_price_callback(query, context, db, db_user, range_unique_id)
        
        # Delete a range
        elif callback_data.startswith("delete_range_"):
            range_unique_id = callback_data.replace("delete_range_", "")
            await admin_delete_range_callback(query, context, db, db_user, range_unique_id)
        
        # View SMS numbers for a range
        elif callback_data.startswith("view_numbers_"):
            range_unique_id = callback_data[len("view_numbers_"):]
            await view_sms_numbers_callback(query, context, db, db_user, range_unique_id)
        
        # Pagination callbacks
        elif callback_data.startswith("sms_page_"):
            page = int(callback_data.split("_")[2])
            await view_sms_ranges_callback(query, context, db, db_user, page=page)
        
        # Make user admin callback
        elif callback_data.startswith("make_admin_"):
            user_id = int(callback_data.split("_")[2])
            await make_admin_callback(query, context, db, db_user, user_id)
        
        # Remove admin callback
        elif callback_data.startswith("remove_admin_"):
            user_id = int(callback_data.split("_")[2])
            await remove_admin_callback(query, context, db, db_user, user_id)
        
        # Ban user callback
        elif callback_data.startswith("ban_user_"):
            user_id = int(callback_data.split("_")[2])
            await ban_user_callback(query, context, db, db_user, user_id)
        
        # Unban user callback
        elif callback_data.startswith("unban_user_"):
            user_id = int(callback_data.split("_")[2])
            await unban_user_callback(query, context, db, db_user, user_id)
        
        # Add balance callback
        elif callback_data.startswith("select_add_balance_"):
            user_id = int(callback_data.split("_")[3])
            await add_balance_prompt_callback(query, context, db, db_user, user_id)
        
        # Deduct balance callback
        elif callback_data.startswith("select_deduct_balance_"):
            user_id = int(callback_data.split("_")[3])
            await deduct_balance_prompt_callback(query, context, db, db_user, user_id)
        
        # Set price for specific range callback
        elif callback_data.startswith("set_price_for_range_"):
            range_id = callback_data[len("set_price_for_range_"):]
            await set_price_for_range_callback(query, context, db, db_user, range_id)
        
        # Select range for price setting
        elif callback_data == "select_range_for_price":
            await select_range_for_price_callback(query, context, db, db_user)
        
        # Add balance by user ID
        elif callback_data == "add_balance_by_id":
            await add_balance_by_id_prompt_callback(query, context, db, db_user)
        
        # Deduct balance by user ID
        elif callback_data == "deduct_balance_by_id":
            await deduct_balance_by_id_prompt_callback(query, context, db, db_user)
        
        # Approve recharge callback
        elif callback_data.startswith("approve_recharge_"):
            request_id = int(callback_data.split("_")[2])
            await approve_recharge_callback(query, context, db, db_user, request_id)
        
        # Reject recharge callback
        elif callback_data.startswith("reject_recharge_"):
            request_id = int(callback_data.split("_")[2])
            await reject_recharge_callback(query, context, db, db_user, request_id)
        
        # Check SMS callback (when user clicks "Check for SMS" button)
        elif callback_data.startswith("check_sms_"):
            range_unique_id = callback_data[len("check_sms_"):]
            await check_sms_callback(query, context, db, db_user, range_unique_id)
        
        # Search SMS for specific number callback
        elif callback_data.startswith("search_sms_"):
            phone_number = callback_data[len("search_sms_"):]
            await search_sms_callback(query, context, db, db_user, phone_number)
        
        # Retry SMS search callback
        elif callback_data.startswith("retry_sms_"):
            await retry_sms_callback(query, context, db, db_user)
    finally:
        db.close()


async def view_sms_ranges_callback(query, context, db, db_user, page=1):
    """Show SMS ranges to the user from database."""
    log_access(db, db_user, f"view_sms_ranges_page_{page}")
    
    # Get all ranges from database with pagination
    all_ranges = get_all_ranges(db)
    
    if not all_ranges:
        error_msg = (
            "❌ No SMS ranges available.\n\n"
            "The administrator needs to upload ranges via CSV file.\n\n"
            "Please contact the administrator or try again later."
        )
        await query.edit_message_text(error_msg)
        logger.info(f"No SMS ranges in database for user {db_user.telegram_id}")
        return
    
    # Implement pagination
    items_per_page = 10
    total_items = len(all_ranges)
    total_pages = (total_items + items_per_page - 1) // items_per_page
    start_idx = (page - 1) * items_per_page
    end_idx = start_idx + items_per_page
    ranges = all_ranges[start_idx:end_idx]
    
    # Create message header with page info
    page_info = f"Page {page}/{total_pages}" if total_pages > 1 else ""
    message = f"📱 Available SMS Ranges ({total_items} total)\n"
    if page_info:
        message += f"{page_info}\n"
    message += "\nSelect a range to view details:"
    
    # Create keyboard with one button per range
    keyboard = []
    
    for range_obj, number_count in ranges:
        # Format button text with range name and number count
        button_text = f"{range_obj.name} ({number_count} numbers)"
        
        # Truncate if too long
        if len(button_text) > MAX_BUTTON_TEXT_LENGTH:
            button_text = button_text[:MAX_BUTTON_TEXT_LENGTH-3] + "..."
        
        keyboard.append([InlineKeyboardButton(
            button_text,
            callback_data=f"range_{range_obj.unique_id}"
        )])
    
    # Add navigation buttons
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Previous", callback_data=f"view_ranges_{page-1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("➡️ Next", callback_data=f"view_ranges_{page+1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    
    # Add back to main menu button
    keyboard.append([InlineKeyboardButton("🏠 Main Menu", callback_data="main_menu")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(message, reply_markup=reply_markup)
async def view_sms_range_detail_callback(query, context, db, db_user, range_unique_id):
    """Show detailed information about a specific SMS range."""
    log_access(db, db_user, f"view_sms_range_detail_{range_unique_id}")
    
    # Get range from database
    range_obj = get_range_by_unique_id(db, range_unique_id)
    
    if not range_obj:
        await query.answer("❌ Range not found.")
        return
    
    # Get total and available number counts
    total_numbers = db.query(PhoneNumber).filter_by(range_id=range_obj.id).count()
    available_numbers = get_available_numbers_for_range(db, range_unique_id, limit=1000)
    available_count = len(available_numbers)
    
    # Get price for this range
    price = get_price_for_range(db, range_unique_id)
    
    # Format the detailed message with HTML
    message = "📱 <b>SMS Range Details</b>\n\n"
    message += f"<b>Name:</b> {escape_html(range_obj.name)}\n"
    message += f"<b>Unique ID:</b> <code>{range_unique_id}</code>\n\n"
    message += f"📊 <b>Statistics:</b>\n"
    message += f"  • Total Numbers: {total_numbers}\n"
    message += f"  • Available: {available_count}\n"
    message += f"  • Held: {total_numbers - available_count}\n\n"
    message += f"💵 <b>Price per SMS:</b> ${price:.2f}\n"
    
    # Add buttons for actions and navigation
    keyboard = []
    
    # If user is admin, show admin options
    if is_user_admin(db, db_user.telegram_id):
        keyboard.append([InlineKeyboardButton("💰 Set Price", callback_data=f"set_price_{range_unique_id}")])
        keyboard.append([InlineKeyboardButton("🗑️ Delete Range", callback_data=f"delete_range_{range_unique_id}")])
    
    keyboard.append([InlineKeyboardButton("📞 Request Numbers", callback_data=f"view_numbers_{range_unique_id}")])
    keyboard.append([InlineKeyboardButton("⬅️ Back to Ranges", callback_data="view_sms_ranges")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def view_sms_numbers_callback(query, context, db, db_user, range_unique_id):
    """Show SMS numbers for a specific range."""
    log_access(db, db_user, f"view_sms_numbers_{range_unique_id}")
    
    # Clean up expired holds
    cleanup_expired_holds(db)
    
    # Get range from database
    range_obj = get_range_by_unique_id(db, range_unique_id)
    if not range_obj:
        await query.answer("❌ Range not found.")
        return
    
    # Get price for this range
    price = get_price_for_range(db, range_unique_id)
    
    # Check if user has sufficient balance
    if db_user.balance < price:
        await query.edit_message_text(
            f"❌ Insufficient balance!\n\n"
            f"💰 Your Balance: ${db_user.balance:.2f}\n"
            f"💵 Required: ${price:.2f}\n\n"
            "Please recharge your balance to continue.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("💰 Recharge", callback_data="recharge_request"),
                InlineKeyboardButton("⬅️ Back", callback_data=f"range_{range_unique_id}")
            ]])
        )
        return
    
    # Get available numbers from database (not held)
    available_numbers = get_available_numbers_for_range(db, range_unique_id, limit=SMS_FETCH_COUNT)
    
    if not available_numbers:
        await query.edit_message_text(
            "❌ No available numbers in this range!\n\n"
            "All numbers are currently held by users.\n"
            "Please try another range or try again later.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("⬅️ Back", callback_data=f"range_{range_unique_id}")
            ]])
        )
        return
    
    # Check if we have enough numbers
    if len(available_numbers) < SMS_DISPLAY_COUNT:
        await query.edit_message_text(
            f"❌ Not enough available numbers in this range!\n\n"
            f"Available: {len(available_numbers)} numbers\n"
            f"Required: {SMS_DISPLAY_COUNT} numbers\n\n"
            "Please try another range or try again later.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("⬅️ Back", callback_data=f"range_{range_unique_id}")
            ]])
            )
        return
    
    # Randomly select SMS_DISPLAY_COUNT numbers
    selected_numbers = random.sample(available_numbers, SMS_DISPLAY_COUNT)
    
    # Create temporary holds for these numbers
    phone_number_ids = [num.id for num in selected_numbers]
    create_number_holds(db, db_user, phone_number_ids, range_unique_id)
    
    # Store selected numbers in context for SMS checking
    if 'selected_numbers' not in context.user_data:
        context.user_data['selected_numbers'] = {}
    context.user_data['selected_numbers'][range_unique_id] = [num.number for num in selected_numbers]
    context.user_data['current_range_id'] = range_unique_id
    
    # Format message with numbers
    message = f"📱 <b>{escape_html(range_obj.name)}</b>\n\n"
    message += f"💵 Price: ${price:.2f} per SMS\n"
    message += f"💰 Your Balance: ${db_user.balance:.2f}\n\n"
    message += f"📞 <b>Your {SMS_DISPLAY_COUNT} Numbers:</b>\n\n"
    
    for i, phone_num in enumerate(selected_numbers, 1):
        message += f"{i}. <code>{phone_num.number}</code>\n"
    
    message += "\n⏰ <b>Numbers are temporarily held for you.</b>\n"
    message += "They will be released after 5 minutes of first search.\n\n"
    message += "Use the buttons below to check for SMS:"
    
    # Create keyboard with check buttons
    keyboard = [
        [InlineKeyboardButton("🔍 Check for SMS", callback_data=f"check_sms_{range_unique_id}")],
        [InlineKeyboardButton("🔄 Get New Numbers", callback_data=f"view_numbers_{range_unique_id}")],
        [InlineKeyboardButton("⬅️ Back to Range", callback_data=f"range_{range_unique_id}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def about_callback(query, context, db, db_user):
    """Show about information."""
    log_access(db, db_user, "about")
    
    message = (
        "ℹ️ About ARGSMS Bot\n\n"
        "This bot provides access to SMS range information.\n"
        "You can view available SMS ranges and navigate through pages.\n\n"
        "Commands:\n"
        "/start - Show main menu\n"
        "/admin - Admin panel (admin only)\n"
    )
    
    keyboard = [[InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_to_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def back_to_main_callback(query, context, db, db_user):
    """Return to main menu."""
    keyboard = [
        [InlineKeyboardButton("📱 View SMS Ranges", callback_data="view_sms_ranges")],
        [InlineKeyboardButton("👤 My Profile", callback_data="user_profile")],
        [InlineKeyboardButton("💰 Recharge Balance", callback_data="recharge_request")],
        [InlineKeyboardButton("ℹ️ About", callback_data="about")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    welcome_message = (
        f"👋 Welcome to ARGSMS Bot!\n\n"
        f"💰 Balance: ${db_user.balance:.2f}\n\n"
        "This bot allows you to view available SMS ranges.\n"
        "Use the menu below to navigate:"
    )
    
    await query.edit_message_text(welcome_message, reply_markup=reply_markup)


async def admin_list_users_callback(query, context, db, db_user):
    """List all users (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_list_users")
    
    # Get all users
    users = db.query(User).order_by(User.created_at.desc()).limit(20).all()
    
    message = "👥 User List (Last 20)\n\n"
    for user in users:
        admin_badge = "🔑" if user.is_admin else "👤"
        ban_badge = "🚫" if user.is_banned else ""
        username_str = f"@{user.username}" if user.username else "N/A"
        message += f"{admin_badge}{ban_badge} ID: {user.telegram_id} | {username_str}\n"
        message += f"   Balance: ${user.balance:.2f} | Joined: {user.created_at.strftime('%Y-%m-%d %H:%M')}\n\n"
    
    keyboard = [
        [InlineKeyboardButton("🔑 Manage Admins", callback_data="admin_manage_admins")],
        [InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def admin_manage_admins_callback(query, context, db, db_user):
    """Manage admin users (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_manage_admins")
    
    # Get all users
    users = db.query(User).order_by(User.created_at.desc()).limit(10).all()
    
    message = "🔑 Manage Admins\n\n"
    message += "Select a user to toggle admin status:\n\n"
    
    keyboard = []
    for user in users:
        admin_badge = "🔑" if user.is_admin else "👤"
        username_str = f"@{user.username}" if user.username else f"ID:{user.telegram_id}"
        button_text = f"{admin_badge} {username_str}"
        
        if user.is_admin:
            callback = f"remove_admin_{user.id}"
        else:
            callback = f"make_admin_{user.id}"
        
        keyboard.append([InlineKeyboardButton(button_text, callback_data=callback)])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def admin_view_stats_callback(query, context, db, db_user):
    """View statistics (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_view_stats")
    
    # Get statistics
    from sqlalchemy import func
    
    total_users = db.query(User).count()
    admin_users = db.query(User).filter_by(is_admin=True).count()
    banned_users = db.query(User).filter_by(is_banned=True).count()
    
    # Calculate total balance in system
    total_balance = db.query(func.sum(User.balance)).scalar() or 0.0
    total_spent = db.query(func.sum(User.total_spent)).scalar() or 0.0
    total_sms = db.query(func.sum(User.total_sms_received)).scalar() or 0
    
    # Number holds stats
    total_holds = db.query(NumberHold).count()
    permanent_holds = db.query(NumberHold).filter_by(is_permanent=True).count()
    
    today = datetime.utcnow().date()
    today_logs = db.query(AccessLog).filter(
        func.date(AccessLog.timestamp) == today
    ).count()
    
    message = (
        "📊 Bot Statistics\n\n"
        f"👥 Total Users: {total_users}\n"
        f"🔑 Admin Users: {admin_users}\n"
        f"🚫 Banned Users: {banned_users}\n\n"
        f"💰 Total Balance in System: ${total_balance:.2f}\n"
        f"💸 Total Spent: ${total_spent:.2f}\n"
        f"📨 Total SMS Received: {total_sms}\n\n"
        f"🔒 Number Holds: {total_holds}\n"
        f"🔐 Permanent Holds: {permanent_holds}\n\n"
        f"📈 Today's Actions: {today_logs}\n"
    )
    
    keyboard = [[InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def admin_number_holds_callback(query, context, db, db_user):
    """Show number holds analysis (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_number_holds")
    
    # Get statistics
    from sqlalchemy import func, Integer
    
    # Total holds
    total_holds = db.query(NumberHold).count()
    permanent_holds = db.query(NumberHold).filter_by(is_permanent=True).count()
    temporary_holds = total_holds - permanent_holds
    
    # Get holds by user
    holds_by_user = db.query(
        User.telegram_id,
        User.username,
        func.count(NumberHold.id).label('hold_count'),
        func.sum(func.cast(NumberHold.is_permanent, Integer)).label('permanent_count')
    ).join(NumberHold).group_by(User.id).order_by(func.count(NumberHold.id).desc()).limit(10).all()
    
    # Get holds by range
    holds_by_range = db.query(
        NumberHold.range_id,
        func.count(NumberHold.id).label('hold_count')
    ).group_by(NumberHold.range_id).order_by(func.count(NumberHold.id).desc()).limit(10).all()
    
    # Check for expired temporary holds
    now = datetime.utcnow()
    expired_holds = db.query(NumberHold).filter(
        NumberHold.is_permanent == False,
        NumberHold.first_retry_time.isnot(None),
        NumberHold.first_retry_time < now - timedelta(minutes=5)
    ).count()
    
    message = (
        "🔒 <b>Number Holds Analysis</b>\n\n"
        f"📊 <b>Overview:</b>\n"
        f"Total Holds: {total_holds}\n"
        f"├─ Permanent: {permanent_holds}\n"
        f"├─ Temporary: {temporary_holds}\n"
        f"└─ Expired (ready to release): {expired_holds}\n\n"
    )
    
    if holds_by_user:
        message += "👥 <b>Top Users by Holds:</b>\n"
        for telegram_id, username, count, perm_count in holds_by_user:
            username_str = f"@{username}" if username else f"ID:{telegram_id}"
            message += f"• {username_str}: {count} holds ({perm_count or 0} permanent)\n"
        message += "\n"
    
    if holds_by_range:
        message += "📱 <b>Top Ranges by Holds:</b>\n"
        for range_id, count in holds_by_range[:5]:
            message += f"• {range_id}: {count} numbers\n"
        message += "\n"
    
    message += "💡 Click 'Export Report' to download detailed Excel report"
    
    keyboard = [
        [InlineKeyboardButton("📥 Export Report", callback_data="admin_export_holds")],
        [InlineKeyboardButton("🔄 Cleanup Expired Holds", callback_data="admin_cleanup_holds")],
        [InlineKeyboardButton("🔓 Release All Temporary Holds", callback_data="admin_release_all_holds")],
        [InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_export_holds_callback(query, context, db, db_user):
    """Export number holds report as Excel file (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    await query.answer("⏳ Generating report...", show_alert=False)
    log_access(db, db_user, "admin_export_holds")
    
    # Get all holds with user information
    holds = db.query(NumberHold, User).join(User).order_by(NumberHold.hold_start_time.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Number Holds Report"
    
    # Define headers
    headers = [
        "User ID", "Username", "Phone Number", "Range ID",
        "Hold Type", "Hold Start", "First Retry", "Status", "Time Info"
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    now = datetime.utcnow()
    for row_idx, (hold, user) in enumerate(holds, start=2):
        # Calculate status and time info
        if hold.is_permanent:
            status = "Permanent"
            time_info = "N/A"
        elif hold.first_retry_time:
            expire_time = hold.first_retry_time + timedelta(minutes=5)
            if now > expire_time:
                status = "Expired"
                time_info = f"Expired {int((now - expire_time).total_seconds() / 60)} min ago"
            else:
                status = "Active"
                remaining = int((expire_time - now).total_seconds() / 60)
                time_info = f"{remaining} min remaining"
        else:
            status = "Active"
            time_info = "Not yet retried"
        
        # Write row data
        ws.cell(row=row_idx, column=1, value=user.telegram_id)
        ws.cell(row=row_idx, column=2, value=user.username or "N/A")
        ws.cell(row=row_idx, column=3, value=hold.phone_number)
        ws.cell(row=row_idx, column=4, value=hold.range_id)
        ws.cell(row=row_idx, column=5, value="Permanent" if hold.is_permanent else "Temporary")
        ws.cell(row=row_idx, column=6, value=hold.hold_start_time.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=7, value=hold.first_retry_time.strftime('%Y-%m-%d %H:%M:%S') if hold.first_retry_time else "N/A")
        ws.cell(row=row_idx, column=8, value=status)
        ws.cell(row=row_idx, column=9, value=time_info)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "Number Holds Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    total_holds = len(holds)
    permanent_count = sum(1 for hold, _ in holds if hold.is_permanent)
    temporary_count = total_holds - permanent_count
    
    expired_count = 0
    active_temp_count = 0
    for hold, _ in holds:
        if not hold.is_permanent:
            if hold.first_retry_time:
                expire_time = hold.first_retry_time + timedelta(minutes=5)
                if now > expire_time:
                    expired_count += 1
                else:
                    active_temp_count += 1
    
    summary_data = [
        ("Total Holds", total_holds),
        ("Permanent Holds", permanent_count),
        ("Temporary Holds", temporary_count),
        ("Active Temporary", active_temp_count),
        ("Expired (Ready to Release)", expired_count),
        ("Report Generated", now.strftime('%Y-%m-%d %H:%M:%S UTC'))
    ]
    
    for idx, (label, value) in enumerate(summary_data, start=3):
        summary_ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=idx, column=2, value=value)
    
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 30
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Send file to admin
    filename = f"number_holds_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    await context.bot.send_document(
        chat_id=query.message.chat_id,
        document=excel_file,
        filename=filename,
        caption=f"📊 Number Holds Report\n\nTotal Records: {total_holds}\nGenerated: {now.strftime('%Y-%m-%d %H:%M:%S UTC')}"
    )
    
    # Return to number holds view
    await admin_number_holds_callback(query, context, db, db_user)


async def admin_cleanup_holds_callback(query, context, db, db_user):
    """Manually cleanup expired holds (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_cleanup_holds")
    
    # Cleanup expired holds
    cleaned = cleanup_expired_holds(db)
    
    await query.answer(f"✅ Cleaned up {cleaned} expired holds", show_alert=True)
    
    # Refresh the number holds view
    await admin_number_holds_callback(query, context, db, db_user)


async def admin_release_all_holds_callback(query, context, db, db_user):
    """Release all temporary holds (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_release_all_holds")
    
    # Count temporary holds before deletion
    temp_holds_count = db.query(NumberHold).filter(
        NumberHold.is_permanent == False
    ).count()
    
    # Delete all temporary holds (keep permanent ones)
    db.query(NumberHold).filter(
        NumberHold.is_permanent == False
    ).delete()
    db.commit()
    
    await query.answer(f"✅ Released {temp_holds_count} temporary holds", show_alert=True)
    
    # Refresh the number holds view
    await admin_number_holds_callback(query, context, db, db_user)


async def make_admin_callback(query, context, db, db_user, target_user_id):
    """Make a user admin (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if target_user:
        target_user.is_admin = True
        db.commit()
        log_access(db, db_user, f"make_admin_{target_user.telegram_id}")
        await query.answer(f"✅ User {target_user.telegram_id} is now an admin")
    else:
        await query.answer("❌ User not found")
    
    # Refresh the manage admins view
    await admin_manage_admins_callback(query, context, db, db_user)


async def remove_admin_callback(query, context, db, db_user, target_user_id):
    """Remove admin status from a user (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    # Prevent removing own admin status
    if target_user_id == db_user.id:
        await query.answer("❌ You cannot remove your own admin status")
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if target_user:
        target_user.is_admin = False
        db.commit()
        log_access(db, db_user, f"remove_admin_{target_user.telegram_id}")
        await query.answer(f"✅ Removed admin status from user {target_user.telegram_id}")
    else:
        await query.answer("❌ User not found")
    
    # Refresh the manage admins view
    await admin_manage_admins_callback(query, context, db, db_user)


async def admin_back_callback(query, context, db, db_user):
    """Return to admin panel."""
    user = query.from_user
    
    # Check if user is admin
    if not is_user_admin(db, user.id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    keyboard = [
        [InlineKeyboardButton("👥 List Users", callback_data="admin_list_users")],
        [InlineKeyboardButton("🔑 Manage Admins", callback_data="admin_manage_admins")],
        [InlineKeyboardButton("🚫 Ban/Unban Users", callback_data="admin_manage_bans")],
        [InlineKeyboardButton("💰 Manage Balance", callback_data="admin_manage_balance")],
        [InlineKeyboardButton("💳 Recharge Requests", callback_data="admin_recharge_requests")],
        [InlineKeyboardButton("💵 Set Price Ranges", callback_data="admin_price_ranges")],
        [InlineKeyboardButton("📊 View Stats", callback_data="admin_view_stats")],
        [InlineKeyboardButton("🔒 Number Holds Report", callback_data="admin_number_holds")],
        [InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_to_main")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    admin_message = (
        "🔐 Admin Panel\n\n"
        "Welcome to the admin panel. Select an option:"
    )
    
    await query.edit_message_text(admin_message, reply_markup=reply_markup)


async def user_profile_callback(query, context, db, db_user):
    """Show user profile with stats."""
    log_access(db, db_user, "view_profile")
    
    # Get user's held numbers count
    held_numbers = db.query(NumberHold).filter_by(user_id=db_user.id).count()
    permanent_holds = db.query(NumberHold).filter_by(user_id=db_user.id, is_permanent=True).count()
    
    message = (
        "👤 <b>Your Profile</b>\n\n"
        f"🆔 <b>User ID:</b> <code>{db_user.telegram_id}</code>\n"
        f"👤 <b>Username:</b> @{db_user.username if db_user.username else 'N/A'}\n"
        f"💰 <b>Balance:</b> ${db_user.balance:.2f}\n"
        f"💸 <b>Total Spent:</b> ${db_user.total_spent:.2f}\n"
        f"📨 <b>Total SMS Received:</b> {db_user.total_sms_received}\n"
        f"🔒 <b>Numbers Held:</b> {held_numbers} ({permanent_holds} permanent)\n"
        f"📅 <b>Joined:</b> {db_user.created_at.strftime('%Y-%m-%d %H:%M')}\n"
    )
    
    keyboard = [[InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_to_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def recharge_request_callback(query, context, db, db_user):
    """Handle recharge request from user."""
    log_access(db, db_user, "recharge_request")
    
    message = (
        "💰 <b>Recharge Balance</b>\n\n"
        "To recharge your balance, please contact an administrator.\n\n"
    )
    
    # Add admin username if configured
    if ADMIN_USERNAME:
        message += f"📞 <b>Contact Admin:</b> @{ADMIN_USERNAME}\n\n"
    
    message += (
        "📝 Send a message to the admin with:\n"
        "• Your User ID\n"
        "• Amount you want to recharge\n"
        "• Payment proof (if required)\n\n"
        f"Your User ID: <code>{db_user.telegram_id}</code>\n"
    )
    
    keyboard = [[InlineKeyboardButton("⬅️ Back to Main Menu", callback_data="back_to_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_manage_bans_callback(query, context, db, db_user):
    """Manage banned users (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_manage_bans")
    
    # Get all users
    users = db.query(User).order_by(User.created_at.desc()).limit(10).all()
    
    message = "🚫 Ban/Unban Users\n\n"
    message += "Select a user to toggle ban status:\n\n"
    
    keyboard = []
    for user in users:
        ban_badge = "🚫" if user.is_banned else "✅"
        username_str = f"@{user.username}" if user.username else f"ID:{user.telegram_id}"
        button_text = f"{ban_badge} {username_str}"
        
        if user.is_banned:
            callback = f"unban_user_{user.id}"
        else:
            callback = f"ban_user_{user.id}"
        
        keyboard.append([InlineKeyboardButton(button_text, callback_data=callback)])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def ban_user_callback(query, context, db, db_user, target_user_id):
    """Ban a user (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if target_user:
        # Don't allow banning admins
        if target_user.is_admin:
            await query.answer("❌ Cannot ban an admin user", show_alert=True)
            return
        
        target_user.is_banned = True
        db.commit()
        log_access(db, db_user, f"ban_user_{target_user.telegram_id}")
        await query.answer(f"✅ User {target_user.telegram_id} has been banned")
    else:
        await query.answer("❌ User not found")
    
    # Refresh the manage bans view
    await admin_manage_bans_callback(query, context, db, db_user)


async def unban_user_callback(query, context, db, db_user, target_user_id):
    """Unban a user (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if target_user:
        target_user.is_banned = False
        db.commit()
        log_access(db, db_user, f"unban_user_{target_user.telegram_id}")
        await query.answer(f"✅ User {target_user.telegram_id} has been unbanned")
    else:
        await query.answer("❌ User not found")
    
    # Refresh the manage bans view
    await admin_manage_bans_callback(query, context, db, db_user)


async def admin_manage_balance_callback(query, context, db, db_user):
    """Manage user balances (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_manage_balance")
    
    # Get all users
    users = db.query(User).order_by(User.created_at.desc()).limit(10).all()
    
    message = "💰 Manage User Balance\n\n"
    message += "Select a user to manage their balance:\n\n"
    
    keyboard = []
    for user in users:
        username_str = f"@{user.username}" if user.username else f"ID:{user.telegram_id}"
        button_text = f"{username_str} - ${user.balance:.2f}"
        keyboard.append([
            InlineKeyboardButton(f"➕ {button_text}", callback_data=f"select_add_balance_{user.id}"),
            InlineKeyboardButton(f"➖", callback_data=f"select_deduct_balance_{user.id}")
        ])
    
    # Add options to manage by user ID
    keyboard.append([InlineKeyboardButton("🆔 Add Balance by User ID", callback_data="add_balance_by_id")])
    keyboard.append([InlineKeyboardButton("🆔 Deduct Balance by User ID", callback_data="deduct_balance_by_id")])
    keyboard.append([InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup)


async def add_balance_prompt_callback(query, context, db, db_user, target_user_id):
    """Prompt for balance addition (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if not target_user:
        await query.answer("❌ User not found", show_alert=True)
        return
    
    # Store target user ID in user_data for next message
    context.user_data['admin_action'] = 'add_balance'
    context.user_data['target_user_id'] = target_user_id
    
    username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
    message = (
        f"💰 <b>Add Balance to {username_str}</b>\n\n"
        f"Current balance: ${target_user.balance:.2f}\n\n"
        "Please send the amount to add (e.g., 100 or 50.5):"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="admin_manage_balance")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def deduct_balance_prompt_callback(query, context, db, db_user, target_user_id):
    """Prompt for balance deduction (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    target_user = db.query(User).filter_by(id=target_user_id).first()
    if not target_user:
        await query.answer("❌ User not found", show_alert=True)
        return
    
    # Store target user ID in user_data for next message
    context.user_data['admin_action'] = 'deduct_balance'
    context.user_data['target_user_id'] = target_user_id
    
    username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
    message = (
        f"💰 <b>Deduct Balance from {username_str}</b>\n\n"
        f"Current balance: ${target_user.balance:.2f}\n\n"
        "Please send the amount to deduct (e.g., 10 or 5.5):"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="admin_manage_balance")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_recharge_requests_callback(query, context, db, db_user):
    """View and manage recharge requests (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_recharge_requests")
    
    # Get pending recharge requests
    pending_requests = db.query(RechargeRequest).filter_by(status='pending').order_by(RechargeRequest.created_at.desc()).limit(10).all()
    
    message = "💳 <b>Recharge Requests</b>\n\n"
    
    if not pending_requests:
        message += "No pending recharge requests.\n\n"
        message += "Users can request recharges from the main menu.\n"
        message += "Contact information will be shown to them."
        keyboard = [[InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]]
    else:
        message += f"Pending requests: {len(pending_requests)}\n\n"
        
        for req in pending_requests:
            req_user = db.query(User).filter_by(id=req.user_id).first()
            username_str = f"@{req_user.username}" if req_user.username else f"ID:{req_user.telegram_id}"
            message += f"📝 {username_str} - ${req.amount:.2f}\n"
            message += f"   Requested: {req.created_at.strftime('%Y-%m-%d %H:%M')}\n\n"
        
        message += "\n💡 Process recharges manually by using 'Manage Balance' option."
        keyboard = [
            [InlineKeyboardButton("💰 Manage Balance", callback_data="admin_manage_balance")],
            [InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]
        ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def approve_recharge_callback(query, context, db, db_user, request_id):
    """Approve a recharge request (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    # This is handled via command
    await query.answer("💡 Use /approverecharge <request_id> command", show_alert=True)


async def reject_recharge_callback(query, context, db, db_user, request_id):
    """Reject a recharge request (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    # This is handled via command
    await query.answer("💡 Use /rejectrecharge <request_id> command", show_alert=True)


async def admin_price_ranges_callback(query, context, db, db_user):
    """Manage price ranges (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_price_ranges")
    
    # Get all price ranges
    price_ranges = db.query(PriceRange).order_by(PriceRange.created_at.desc()).limit(10).all()
    
    message = "💵 <b>SMS Price Ranges</b>\n\n"
    
    if not price_ranges:
        message += "No price ranges configured.\n\n"
    else:
        for pr in price_ranges:
            message += f"📍 Range: <code>{pr.range_pattern}</code>\n"
            message += f"   Price: ${pr.price:.2f}\n\n"
    
    message += "Select an option below:"
    
    keyboard = [
        [InlineKeyboardButton("📋 Select from Available Ranges", callback_data="select_range_for_price")],
        [InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def add_balance_by_id_prompt_callback(query, context, db, db_user):
    """Prompt admin to add balance by entering user ID."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    context.user_data['admin_action'] = 'add_balance_by_id_step1'
    
    message = (
        "💰 <b>Add Balance by User ID</b>\n\n"
        "Please send the Telegram User ID (e.g., 123456789):"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="admin_manage_balance")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def deduct_balance_by_id_prompt_callback(query, context, db, db_user):
    """Prompt admin to deduct balance by entering user ID."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    context.user_data['admin_action'] = 'deduct_balance_by_id_step1'
    
    message = (
        "💰 <b>Deduct Balance by User ID</b>\n\n"
        "Please send the Telegram User ID (e.g., 123456789):"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="admin_manage_balance")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def select_range_for_price_callback(query, context, db, db_user):
    """Show available SMS ranges for price setting (admin only)."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "select_range_for_price")
    
    # Get scrapper session and fetch ranges
    scrapper = get_scrapper_session()
    data = scrapper.get_sms_ranges(max_results=20, page=1)
    
    if not data:
        error_msg = (
            "❌ Failed to retrieve SMS ranges.\n\n"
            "This could be due to:\n"
            "• API server is temporarily down\n"
            "• Network connection issue\n"
            "• Authentication failure\n\n"
            "Please try again in a few moments."
        )
        await query.edit_message_text(
            error_msg,
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("⬅️ Back", callback_data="admin_price_ranges")
            ]])
        )
        logger.error(f"Failed to retrieve SMS ranges for admin {db_user.telegram_id}")
        return
    
    # Parse ranges
    ranges = []
    if isinstance(data, dict):
        if 'results' in data:
            ranges = data['results']
        elif 'data' in data:
            ranges = data['data']
        elif 'aaData' in data:
            ranges = data['aaData']
        else:
            ranges = [data]
    elif isinstance(data, list):
        ranges = data
    
    message = "💵 <b>Select Range to Set Price</b>\n\n"
    message += f"Available SMS Ranges ({len(ranges)}):\n"
    message += "Click on a range to set its price.\n\n"
    
    # Create keyboard with range buttons
    keyboard = []
    
    for i, item in enumerate(ranges, 1):
        button_text = ""
        range_id = None
        
        if isinstance(item, dict):
            range_id = item.get('id') or item.get('range_id') or str(i)
            title = item.get('title', '')
            
            if title:
                button_text = f"{title[:MAX_TITLE_LENGTH]}" if len(title) > MAX_TITLE_LENGTH else title
            else:
                info = " - ".join(f"{k}: {v}" for k, v in list(item.items())[:2])
                button_text = info[:MAX_BUTTON_TEXT_LENGTH]
        elif isinstance(item, list):
            range_id = str(i)
            info = " | ".join(str(x) for x in item[:2])
            button_text = info[:MAX_BUTTON_TEXT_LENGTH]
        else:
            range_id = str(i)
            button_text = str(item)[:MAX_BUTTON_TEXT_LENGTH]
        
        # Store range data for later
        if 'price_ranges_data' not in context.chat_data:
            context.chat_data['price_ranges_data'] = {}
        context.chat_data['price_ranges_data'][str(range_id)] = item
        
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f"set_price_for_range_{range_id}")])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="admin_price_ranges")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def set_price_for_range_callback(query, context, db, db_user, range_id):
    """Prompt admin to set price for a specific range."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    # Get range data
    range_data = None
    if 'price_ranges_data' in context.chat_data:
        range_data = context.chat_data['price_ranges_data'].get(str(range_id))
    
    # Extract range title/name
    range_name = range_id
    if range_data:
        if isinstance(range_data, dict):
            range_name = range_data.get('title', range_data.get('name', range_id))
        elif isinstance(range_data, list) and len(range_data) > 0:
            range_name = str(range_data[0])
    
    # Check if price already exists for this range
    existing = db.query(PriceRange).filter_by(range_pattern=range_id).first()
    current_price = existing.price if existing else None
    
    # Store range ID for later
    context.user_data['admin_action'] = 'set_price_for_specific_range'
    context.user_data['selected_range_id'] = range_id
    context.user_data['selected_range_name'] = range_name
    
    message = (
        f"💵 <b>Set Price for Range</b>\n\n"
        f"Range: <code>{escape_html(str(range_name))}</code>\n"
    )
    
    if current_price is not None:
        message += f"Current Price: ${current_price:.2f}\n\n"
    else:
        message += "Current Price: Not set\n\n"
    
    message += "Please send the new price (e.g., 2.5 or 10):"
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="select_range_for_price")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


def is_phone_number(text):
    """Check if the text is a valid phone number."""
    # Remove any whitespace or special characters
    cleaned = re.sub(r'[^\d+]', '', text)
    
    # Check if it's a valid phone number (at least 10 digits, optionally starting with +)
    # Pattern: optional +, then 10-15 digits
    pattern = r'^\+?\d{10,15}$'
    return bool(re.match(pattern, cleaned))


def is_stats_row(row):
    """
    Check if a row from the SMS messages API is a statistics row rather than an actual message.
    
    Stats rows typically have comma-separated values with percentages in the first field,
    e.g., "0.01,0,0,0,0,0,0.01,0,0,100%,NAN%,NAN%,1"
    
    Args:
        row: A list representing a row from the aaData array
        
    Returns:
        True if the row is a stats row, False if it's an actual SMS message
    """
    if not isinstance(row, list) or len(row) < 1:
        return False
    
    first_field = row[0]
    if not isinstance(first_field, str):
        return False
    
    # Stats rows have comma-separated values with percentage signs
    return ',' in first_field and '%' in first_field


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle all text messages - check for admin actions first, then phone search."""
    message = update.message
    if not message or not message.text:
        return
    
    text = message.text.strip()
    db = get_db_session()
    
    try:
        user = get_or_create_user(
            db,
            telegram_id=message.from_user.id,
            username=message.from_user.username or message.from_user.first_name
        )
        
        # Check for admin actions first
        admin_action = context.user_data.get('admin_action')
        if admin_action and is_user_admin(db, user.telegram_id):
            # Handle admin input
            await handle_admin_input(update, context)
            return
        
        # Check if the message is a phone number
        if not is_phone_number(text):
            return
        
        # Handle phone search
        await handle_phone_search(update, context)
    finally:
        db.close()


async def handle_phone_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle phone number search when user sends a phone number."""
    message = update.message
    if not message or not message.text:
        return
    
    text = message.text.strip()
    
    # Extract clean phone number
    phone_number = re.sub(r'[^\d+]', '', text)
    
    # Get database session
    db = get_db_session()
    try:
        # Get or create user
        user = get_or_create_user(
            db,
            telegram_id=message.from_user.id,
            username=message.from_user.username or message.from_user.first_name
        )
        
        # Check if user is banned
        if user.is_banned:
            await message.reply_text(
                "🚫 You have been banned from using this bot.\n"
                "Please contact an administrator if you believe this is an error."
            )
            return
        
        # Check channel membership
        if not await check_channel_membership(update, context, db):
            return
        
        # Check if this number is held by the user
        held_number = db.query(NumberHold).filter_by(
            user_id=user.id,
            phone_number_str=phone_number
        ).first()
        
        if not held_number:
            await message.reply_text(
                "❌ This number is not held by you.\n"
                "Please request numbers from a range first."
            )
            return
        
        # Update first retry time if not set
        if not held_number.first_retry_time:
            update_first_retry_time(db, user, phone_number)
        
        # Log the search
        log_access(db, user, f"SMS search: {phone_number}")
        
        # Send "searching" message
        searching_msg = await message.reply_text(
            f"🔍 Searching SMS messages for <code>{escape_html(phone_number)}</code>...",
            parse_mode='HTML'
        )
        
        try:
            # Get scrapper session
            scrapper = get_scrapper_session()
            
            # Search for SMS messages
            data = scrapper.get_sms_messages(phone_number)
            
            if not data:
                await searching_msg.edit_text(
                    "❌ Failed to retrieve SMS messages. Please try again later."
                )
                return
            
            # Parse response
            total_records = data.get('iTotalRecords', '0')
            if isinstance(total_records, str):
                total_records = int(total_records) if total_records.isdigit() else 0
            
            messages = data.get('aaData', [])
            
            # Filter out the stats row using helper function
            actual_messages = []
            for msg in messages:
                if isinstance(msg, list) and len(msg) >= 6:
                    if not is_stats_row(msg):
                        actual_messages.append(msg)
            
            if not actual_messages:
                # Create keyboard with retry button
                keyboard = [
                    [
                        InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                        InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
                    ]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                await searching_msg.edit_text(
                    f"📭 No SMS messages found for <code>{escape_html(phone_number)}</code>",
                    parse_mode='HTML',
                    reply_markup=reply_markup
                )
                return
            
            # SMS found! Mark as permanent and deduct balance
            was_temporary = not held_number.is_permanent
            price = 0.0
            
            if was_temporary:
                # Get price for this range
                price = get_price_for_range(db, held_number.range_id)
                
                # Deduct balance
                new_balance = deduct_user_balance(
                    db, user, price, 
                    transaction_type='sms_charge',
                    description=f"SMS received on {phone_number}"
                )
                
                if new_balance is None:
                    await searching_msg.edit_text(
                        "❌ Insufficient balance to complete this transaction."
                    )
                    return
                
                # Mark as permanent hold
                mark_number_permanent(db, user, phone_number)
            
            # Format messages
            message_text = f"📱 <b>SMS Messages for {escape_html(phone_number)}</b>\n"
            message_text += f"Found {len(actual_messages)} message(s)\n"
            
            # Show balance deduction info if it was just deducted
            if was_temporary:
                message_text += f"\n💸 Balance deducted: ${price:.2f}\n"
                message_text += f"💰 New balance: ${user.balance:.2f}\n"
            
            message_text += "\n"
            
            for i, msg_data in enumerate(actual_messages, 1):
                time = msg_data[0] if len(msg_data) > 0 else "N/A"
                sender_id = msg_data[3] if len(msg_data) > 3 else "N/A"
                sms_body = msg_data[5] if len(msg_data) > 5 else "N/A"
                
                # Clean up None values
                if sender_id is None:
                    sender_id = "Unknown"
                if sms_body is None:
                    sms_body = "(empty)"
                
                # Escape HTML in the content
                time_escaped = escape_html(str(time))
                sender_escaped = escape_html(str(sender_id))
                body_escaped = escape_html(str(sms_body))
                
                message_text += f"<b>Message {i}:</b>\n"
                message_text += f"🕒 <b>Time:</b> {time_escaped}\n"
                message_text += f"📨 <b>Sender:</b> {sender_escaped}\n"
                message_text += f"💬 <b>Message:</b>\n<pre>{body_escaped}</pre>\n\n"
                
                # Telegram has a message length limit, so split if needed
                if len(message_text) > MAX_TELEGRAM_MESSAGE_LENGTH:
                    message_text += f"<i>... and {len(actual_messages) - i} more message(s)</i>"
                    break
            
            # Create keyboard with back button
            keyboard = [[InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await searching_msg.edit_text(
                message_text,
                parse_mode='HTML',
                reply_markup=reply_markup
            )
            
        except Exception as e:
            logger.error(f"Error in phone search: {e}")
            await searching_msg.edit_text(
                "❌ An error occurred while searching for SMS messages. Please try again later."
            )
    finally:
        db.close()


async def check_sms_callback(query, context, db, db_user, range_unique_id):
    """Handle 'Check for SMS' button click - show list of held numbers to check."""
    log_access(db, db_user, f"check_sms_{range_unique_id}")
    
    # Get the numbers stored in context for this range
    if 'selected_numbers' not in context.user_data or range_unique_id not in context.user_data['selected_numbers']:
        await query.edit_message_text(
            "❌ No numbers found. Please request numbers again.",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("⬅️ Back", callback_data=f"range_{range_unique_id}")
            ]])
        )
        return
    
    numbers = context.user_data['selected_numbers'][range_unique_id]
    
    # Create message with all numbers and instructions
    message = f"📱 <b>Check SMS for Numbers</b>\n\n"
    message += f"Your {len(numbers)} held numbers:\n\n"
    
    for i, phone_number in enumerate(numbers, 1):
        message += f"{i}. <code>{phone_number}</code>\n"
    
    message += "\n💡 <b>How to check:</b>\n"
    message += "Simply send any phone number from the list above as a message, and I'll search for SMS!\n\n"
    message += "Or click a button below to search for a specific number:"
    
    # Create keyboard with buttons for each number (max 20)
    keyboard = []
    for phone_number in numbers[:20]:  # Limit to 20 buttons
        keyboard.append([InlineKeyboardButton(
            f"🔍 {phone_number}", 
            callback_data=f"search_sms_{phone_number}"
        )])
    
    # Add navigation buttons
    keyboard.append([
        InlineKeyboardButton("🔄 Get New Numbers", callback_data=f"view_numbers_{range_unique_id}"),
        InlineKeyboardButton("⬅️ Back", callback_data=f"range_{range_unique_id}")
    ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        message,
        parse_mode='HTML',
        reply_markup=reply_markup
    )


async def search_sms_callback(query, context, db, db_user, phone_number):
    """Search for SMS messages for a specific phone number (from button click)."""
    # Check if this number is held by the user
    held_number = db.query(NumberHold).filter_by(
        user_id=db_user.id,
        phone_number_str=phone_number
    ).first()
    
    if not held_number:
        await query.answer("❌ This number is not held by you.", show_alert=True)
        return
    
    # Update first retry time if not set
    if not held_number.first_retry_time:
        update_first_retry_time(db, db_user, phone_number)
    
    # Log the search
    log_access(db, db_user, f"SMS search: {phone_number}")
    
    # Show searching message
    await query.edit_message_text(
        f"🔍 Searching SMS messages for <code>{escape_html(phone_number)}</code>...",
        parse_mode='HTML'
    )
    
    try:
        # Get scrapper session
        scrapper = get_scrapper_session()
        
        # Search for SMS messages
        data = scrapper.get_sms_messages(phone_number)
        
        if not data:
            # Create keyboard with retry button
            keyboard = [
                [
                    InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                    InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                "❌ Failed to retrieve SMS messages. Please try again.",
                reply_markup=reply_markup
            )
            return
        
        # Parse response
        total_records = data.get('iTotalRecords', '0')
        if isinstance(total_records, str):
            total_records = int(total_records) if total_records.isdigit() else 0
        
        messages = data.get('aaData', [])
        
        # Filter out the stats row
        actual_messages = []
        for msg in messages:
            if isinstance(msg, list) and len(msg) >= 6:
                if not is_stats_row(msg):
                    actual_messages.append(msg)
        
        if not actual_messages:
            # Create keyboard with retry button
            keyboard = [
                [
                    InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                    InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                f"📭 No SMS messages found for <code>{escape_html(phone_number)}</code>",
                parse_mode='HTML',
                reply_markup=reply_markup
            )
            return
        
        # SMS found! Mark as permanent and deduct balance
        was_temporary = not held_number.is_permanent
        price = 0.0
        
        if was_temporary:
            # Get price for this range
            price = get_price_for_range(db, held_number.range_id)
            
            # Deduct balance
            new_balance = deduct_user_balance(
                db, db_user, price, 
                transaction_type='sms_charge',
                description=f"SMS received on {phone_number}"
            )
            
            if new_balance is None:
                await query.edit_message_text(
                    "❌ Insufficient balance to complete this transaction."
                )
                return
            
            # Mark as permanent hold
            mark_number_permanent(db, db_user, phone_number)
        
        # Format messages
        message_text = f"📱 <b>SMS Messages for {escape_html(phone_number)}</b>\n"
        message_text += f"Found {len(actual_messages)} message(s)\n"
        
        # Show balance deduction info if it was just deducted
        if was_temporary:
            message_text += f"\n💸 Balance deducted: ${price:.2f}\n"
            message_text += f"💰 New balance: ${db_user.balance:.2f}\n"
        
        message_text += "\n"
        
        for i, msg_data in enumerate(actual_messages, 1):
            time = msg_data[0] if len(msg_data) > 0 else "N/A"
            sender_id = msg_data[3] if len(msg_data) > 3 else "N/A"
            sms_body = msg_data[5] if len(msg_data) > 5 else "N/A"
            
            # Clean up None values
            if sender_id is None:
                sender_id = "Unknown"
            if sms_body is None:
                sms_body = "(empty)"
            
            # Escape HTML in the content
            time_escaped = escape_html(str(time))
            sender_escaped = escape_html(str(sender_id))
            body_escaped = escape_html(str(sms_body))
            
            message_text += f"<b>Message {i}:</b>\n"
            message_text += f"🕒 <b>Time:</b> {time_escaped}\n"
            message_text += f"📨 <b>Sender:</b> {sender_escaped}\n"
            message_text += f"💬 <b>Message:</b>\n<pre>{body_escaped}</pre>\n\n"
            
            # Telegram has a message length limit, so split if needed
            if len(message_text) > MAX_TELEGRAM_MESSAGE_LENGTH:
                message_text += f"<i>... and {len(actual_messages) - i} more message(s)</i>"
                break
        
        # Create keyboard with back button
        keyboard = [[InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            message_text,
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        
    except Exception as e:
        logger.error(f"Error in SMS search: {e}")
        # Create keyboard with retry button
        keyboard = [
            [
                InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "❌ An error occurred while searching for SMS messages. Please try again.",
            reply_markup=reply_markup
        )


async def retry_sms_callback(query, context, db, db_user):
    """Handle retry button for SMS search."""
    # Extract phone number from callback data (format: retry_sms_PHONENUMBER)
    phone_number = query.data.replace("retry_sms_", "")
    
    # Check if this number is held by the user
    held_number = db.query(NumberHold).filter_by(
        user_id=db_user.id,
        phone_number_str=phone_number
    ).first()
    
    if not held_number:
        await query.answer("❌ This number is not held by you.", show_alert=True)
        return
    
    # Update first retry time if not set
    if not held_number.first_retry_time:
        update_first_retry_time(db, db_user, phone_number)
    
    # Log the retry
    log_access(db, db_user, f"SMS search retry: {phone_number}")
    
    # Show searching message
    await query.edit_message_text(
        f"🔍 Searching SMS messages for <code>{escape_html(phone_number)}</code>...",
        parse_mode='HTML'
    )
    
    try:
        # Get scrapper session
        scrapper = get_scrapper_session()
        
        # Search for SMS messages
        data = scrapper.get_sms_messages(phone_number)
        
        if not data:
            # Create keyboard with retry button
            keyboard = [
                [
                    InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                    InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                "❌ Failed to retrieve SMS messages. Please try again.",
                reply_markup=reply_markup
            )
            return
        
        # Parse response
        total_records = data.get('iTotalRecords', '0')
        if isinstance(total_records, str):
            total_records = int(total_records) if total_records.isdigit() else 0
        
        messages = data.get('aaData', [])
        
        # Filter out the stats row
        actual_messages = []
        for msg in messages:
            if isinstance(msg, list) and len(msg) >= 6:
                if not is_stats_row(msg):
                    actual_messages.append(msg)
        
        if not actual_messages:
            # Create keyboard with retry button
            keyboard = [
                [
                    InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                    InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
                ]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await query.edit_message_text(
                f"📭 No SMS messages found for <code>{escape_html(phone_number)}</code>",
                parse_mode='HTML',
                reply_markup=reply_markup
            )
            return
        
        # SMS found! Mark as permanent and deduct balance
        was_temporary = not held_number.is_permanent
        price = 0.0
        
        if was_temporary:
            # Get price for this range
            price = get_price_for_range(db, held_number.range_id)
            
            # Deduct balance
            new_balance = deduct_user_balance(
                db, db_user, price, 
                transaction_type='sms_charge',
                description=f"SMS received on {phone_number}"
            )
            
            if new_balance is None:
                await query.edit_message_text(
                    "❌ Insufficient balance to complete this transaction."
                )
                return
            
            # Mark as permanent hold
            mark_number_permanent(db, db_user, phone_number)
        
        # Format messages
        message_text = f"📱 <b>SMS Messages for {escape_html(phone_number)}</b>\n"
        message_text += f"Found {len(actual_messages)} message(s)\n"
        
        # Show balance deduction info if it was just deducted
        if was_temporary:
            message_text += f"\n💸 Balance deducted: ${price:.2f}\n"
            message_text += f"💰 New balance: ${db_user.balance:.2f}\n"
        
        message_text += "\n"
        
        for i, msg_data in enumerate(actual_messages, 1):
            time = msg_data[0] if len(msg_data) > 0 else "N/A"
            sender_id = msg_data[3] if len(msg_data) > 3 else "N/A"
            sms_body = msg_data[5] if len(msg_data) > 5 else "N/A"
            
            # Clean up None values
            if sender_id is None:
                sender_id = "Unknown"
            if sms_body is None:
                sms_body = "(empty)"
            
            # Escape HTML in the content
            time_escaped = escape_html(str(time))
            sender_escaped = escape_html(str(sender_id))
            body_escaped = escape_html(str(sms_body))
            
            message_text += f"<b>Message {i}:</b>\n"
            message_text += f"🕒 <b>Time:</b> {time_escaped}\n"
            message_text += f"📨 <b>Sender:</b> {sender_escaped}\n"
            message_text += f"💬 <b>Message:</b>\n<pre>{body_escaped}</pre>\n\n"
            
            # Telegram has a message length limit
            if len(message_text) > MAX_TELEGRAM_MESSAGE_LENGTH:
                message_text += f"<i>... and {len(actual_messages) - i} more message(s)</i>"
                break
        
        # Create keyboard with back button
        keyboard = [[InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            message_text,
            parse_mode='HTML',
            reply_markup=reply_markup
        )
        
    except Exception as e:
        logger.error(f"Error in SMS retry: {e}")
        
        # Create keyboard with retry button
        keyboard = [
            [
                InlineKeyboardButton("🔄 Retry", callback_data=f"retry_sms_{phone_number}"),
                InlineKeyboardButton("🏠 Main Menu", callback_data="back_to_main")
            ]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "❌ An error occurred while searching for SMS messages. Please try again.",
            reply_markup=reply_markup
        )


async def handle_admin_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle admin text inputs for balance and price operations."""
    message = update.message
    if not message or not message.text:
        return
    
    text = message.text.strip()
    db = get_db_session()
    
    try:
        # Get user
        user = get_or_create_user(
            db,
            telegram_id=message.from_user.id,
            username=message.from_user.username
        )
        
        # Check if user is admin
        if not is_user_admin(db, user.telegram_id):
            return
        
        # Check if there's a pending admin action
        admin_action = context.user_data.get('admin_action')
        
        if admin_action == 'add_balance_by_id_step1':
            # Parse user ID
            try:
                target_telegram_id = int(text)
                
                # Find user
                target_user = db.query(User).filter_by(telegram_id=target_telegram_id).first()
                if not target_user:
                    await message.reply_text(f"❌ User with ID {target_telegram_id} not found.")
                    context.user_data.clear()
                    return
                
                # Ask for amount
                context.user_data['admin_action'] = 'add_balance_by_id_step2'
                context.user_data['target_telegram_id'] = target_telegram_id
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"💰 <b>Add Balance to {username_str}</b>\n\n"
                    f"Current balance: ${target_user.balance:.2f}\n\n"
                    "Please send the amount to add (e.g., 100 or 50.5):",
                    parse_mode='HTML'
                )
                
            except ValueError:
                await message.reply_text("❌ Invalid user ID. Please send a valid number:")
                return
        
        elif admin_action == 'add_balance_by_id_step2':
            # Parse amount
            try:
                amount = float(text)
                if amount <= 0:
                    await message.reply_text("❌ Amount must be positive. Please try again:")
                    return
                
                target_telegram_id = context.user_data.get('target_telegram_id')
                target_user = db.query(User).filter_by(telegram_id=target_telegram_id).first()
                
                if not target_user:
                    await message.reply_text("❌ User not found.")
                    context.user_data.clear()
                    return
                
                # Add balance
                new_balance = add_user_balance(
                    db, target_user, amount,
                    transaction_type='admin_add',
                    description=f"Admin added by {user.telegram_id}"
                )
                
                log_access(db, user, f"add_balance_{target_telegram_id}_{amount}")
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"✅ Successfully added ${amount:.2f} to {username_str}\n"
                    f"New balance: ${new_balance:.2f}"
                )
                
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid amount. Please send a valid number (e.g., 100 or 50.5):")
                return
        
        elif admin_action == 'deduct_balance_by_id_step1':
            # Parse user ID
            try:
                target_telegram_id = int(text)
                
                # Find user
                target_user = db.query(User).filter_by(telegram_id=target_telegram_id).first()
                if not target_user:
                    await message.reply_text(f"❌ User with ID {target_telegram_id} not found.")
                    context.user_data.clear()
                    return
                
                # Ask for amount
                context.user_data['admin_action'] = 'deduct_balance_by_id_step2'
                context.user_data['target_telegram_id'] = target_telegram_id
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"💰 <b>Deduct Balance from {username_str}</b>\n\n"
                    f"Current balance: ${target_user.balance:.2f}\n\n"
                    "Please send the amount to deduct (e.g., 10 or 5.5):",
                    parse_mode='HTML'
                )
                
            except ValueError:
                await message.reply_text("❌ Invalid user ID. Please send a valid number:")
                return
        
        elif admin_action == 'deduct_balance_by_id_step2':
            # Parse amount
            try:
                amount = float(text)
                if amount <= 0:
                    await message.reply_text("❌ Amount must be positive. Please try again:")
                    return
                
                target_telegram_id = context.user_data.get('target_telegram_id')
                target_user = db.query(User).filter_by(telegram_id=target_telegram_id).first()
                
                if not target_user:
                    await message.reply_text("❌ User not found.")
                    context.user_data.clear()
                    return
                
                # Check sufficient balance
                if target_user.balance < amount:
                    await message.reply_text(
                        f"❌ User has insufficient balance.\n"
                        f"Current balance: ${target_user.balance:.2f}\n"
                        f"Requested deduction: ${amount:.2f}"
                    )
                    context.user_data.clear()
                    return
                
                # Deduct balance
                new_balance = deduct_user_balance(
                    db, target_user, amount,
                    transaction_type='admin_deduct',
                    description=f"Admin deducted by {user.telegram_id}"
                )
                
                log_access(db, user, f"deduct_balance_{target_telegram_id}_{amount}")
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"✅ Successfully deducted ${amount:.2f} from {username_str}\n"
                    f"New balance: ${new_balance:.2f}"
                )
                
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid amount. Please send a valid number (e.g., 10 or 5.5):")
                return
        
        elif admin_action == 'add_balance':
            # Parse amount
            try:
                amount = float(text)
                if amount <= 0:
                    await message.reply_text("❌ Amount must be positive. Please try again:")
                    return
                
                target_user_id = context.user_data.get('target_user_id')
                target_user = db.query(User).filter_by(id=target_user_id).first()
                
                if not target_user:
                    await message.reply_text("❌ User not found.")
                    context.user_data.clear()
                    return
                
                # Add balance
                new_balance = add_user_balance(
                    db, target_user, amount,
                    transaction_type='admin_add',
                    description=f"Admin added by {user.telegram_id}"
                )
                
                log_access(db, user, f"add_balance_{target_user.telegram_id}_{amount}")
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"✅ Successfully added ${amount:.2f} to {username_str}\n"
                    f"New balance: ${new_balance:.2f}"
                )
                
                # Clear state
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid amount. Please send a valid number (e.g., 100 or 50.5):")
                return
        
        elif admin_action == 'deduct_balance':
            # Parse amount
            try:
                amount = float(text)
                if amount <= 0:
                    await message.reply_text("❌ Amount must be positive. Please try again:")
                    return
                
                target_user_id = context.user_data.get('target_user_id')
                target_user = db.query(User).filter_by(id=target_user_id).first()
                
                if not target_user:
                    await message.reply_text("❌ User not found.")
                    context.user_data.clear()
                    return
                
                # Check sufficient balance
                if target_user.balance < amount:
                    await message.reply_text(
                        f"❌ User has insufficient balance.\n"
                        f"Current balance: ${target_user.balance:.2f}\n"
                        f"Requested deduction: ${amount:.2f}"
                    )
                    context.user_data.clear()
                    return
                
                # Deduct balance
                new_balance = deduct_user_balance(
                    db, target_user, amount,
                    transaction_type='admin_deduct',
                    description=f"Admin deducted by {user.telegram_id}"
                )
                
                log_access(db, user, f"deduct_balance_{target_user.telegram_id}_{amount}")
                
                username_str = f"@{target_user.username}" if target_user.username else f"ID:{target_user.telegram_id}"
                await message.reply_text(
                    f"✅ Successfully deducted ${amount:.2f} from {username_str}\n"
                    f"New balance: ${new_balance:.2f}"
                )
                
                # Clear state
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid amount. Please send a valid number (e.g., 10 or 5.5):")
                return
        
        elif admin_action == 'set_price_for_specific_range':
            # Parse price for a specific range
            try:
                price = float(text)
                if price <= 0:
                    await message.reply_text("❌ Price must be positive. Please try again:")
                    return
                
                range_id = context.user_data.get('selected_range_id')
                range_name = context.user_data.get('selected_range_name', range_id)
                
                # Check if price already exists
                existing = db.query(PriceRange).filter_by(range_pattern=range_id).first()
                
                if existing:
                    # Update existing
                    existing.price = price
                    db.commit()
                    action = "updated"
                else:
                    # Create new
                    price_range = PriceRange(
                        range_pattern=range_id,
                        price=price,
                        created_by=user.id
                    )
                    db.add(price_range)
                    db.commit()
                    action = "created"
                
                log_access(db, user, f"set_price_{range_id}_{price}")
                
                await message.reply_text(
                    f"✅ Price {action} successfully!\n"
                    f"Range: <code>{escape_html(str(range_name))}</code>\n"
                    f"Price: ${price:.2f}",
                    parse_mode='HTML'
                )
                
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid price. Please send a valid number (e.g., 2.5 or 10):")
                return
        
        elif admin_action == 'set_price_pattern':
            # Store pattern and ask for price
            context.user_data['price_pattern'] = text.lower()
            context.user_data['admin_action'] = 'set_price_amount'
            
            await message.reply_text(
                f"💵 Pattern set: <code>{text.lower()}</code>\n\n"
                "Now send the price (e.g., 2.5 or 10):",
                parse_mode='HTML'
            )
            return
        
        elif admin_action == 'set_price_amount':
            # Parse price
            try:
                price = float(text)
                if price <= 0:
                    await message.reply_text("❌ Price must be positive. Please try again:")
                    return
                
                pattern = context.user_data.get('price_pattern')
                
                # Check if pattern already exists
                existing = db.query(PriceRange).filter_by(range_pattern=pattern).first()
                
                if existing:
                    # Update existing
                    existing.price = price
                    db.commit()
                    action = "updated"
                else:
                    # Create new
                    price_range = PriceRange(
                        range_pattern=pattern,
                        price=price,
                        created_by=user.id
                    )
                    db.add(price_range)
                    db.commit()
                    action = "created"
                
                log_access(db, user, f"set_price_{pattern}_{price}")
                
                await message.reply_text(
                    f"✅ Price range {action} successfully!\n"
                    f"Pattern: <code>{pattern}</code>\n"
                    f"Price: ${price:.2f}",
                    parse_mode='HTML'
                )
                
                # Clear state
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid price. Please send a valid number (e.g., 2.5 or 10):")
                return
        
        elif admin_action == 'set_range_price':
            # Handle price input for a specific range
            try:
                price = float(text)
                if price <= 0:
                    await message.reply_text("❌ Price must be positive. Please try again:")
                    return
                
                range_unique_id = context.user_data.get('range_unique_id')
                range_name = context.user_data.get('range_name', 'Unknown')
                
                # Set the price
                set_range_price(db, range_unique_id, range_name, price, user)
                
                log_access(db, user, f"set_range_price_{range_unique_id}_{price}")
                
                await message.reply_text(
                    f"✅ Price set successfully!\n"
                    f"Range: <b>{escape_html(range_name)}</b>\n"
                    f"New Price: ${price:.2f}",
                    parse_mode='HTML'
                )
                
                context.user_data.clear()
                
            except ValueError:
                await message.reply_text("❌ Invalid price. Please send a valid number (e.g., 1.50 or 5):")
                return
    
    finally:
        db.close()


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads (CSV files)."""
    message = update.message
    if not message or not message.document:
        return
    
    db = get_db_session()
    
    try:
        user = get_or_create_user(
            db,
            telegram_id=message.from_user.id,
            username=message.from_user.username
        )
        
        # Check if user is admin
        if not is_user_admin(db, user.telegram_id):
            await message.reply_text("❌ Admin access required to upload files.")
            return
        
        # Check if waiting for CSV upload
        admin_action = context.user_data.get('admin_action')
        if admin_action != 'upload_csv':
            return
        
        document = message.document
        
        # Check file type
        if not document.file_name or not document.file_name.lower().endswith('.csv'):
            await message.reply_text("❌ Please upload a CSV file (.csv extension)")
            return
        
        # Download the file
        file = await context.bot.get_file(document.file_id)
        import tempfile
        import os
        
        # Create temporary file
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.csv', delete=False) as temp_file:
            temp_path = temp_file.name
            await file.download_to_drive(temp_path)
        
        try:
            # Send processing message
            processing_msg = await message.reply_text("📤 Processing CSV file...")
            
            # Import CSV data
            success_count, error_count, errors = import_csv_data(db, temp_path)
            
            # Format result message
            result_msg = f"✅ <b>CSV Import Complete</b>\n\n"
            result_msg += f"✔️ Successfully imported: {success_count} numbers\n"
            
            if error_count > 0:
                result_msg += f"❌ Errors: {error_count}\n\n"
                
                if errors:
                    result_msg += "<b>Error Details:</b>\n"
                    # Show first 10 errors
                    for error in errors[:10]:
                        result_msg += f"• {error}\n"
                    
                    if len(errors) > 10:
                        result_msg += f"\n... and {len(errors) - 10} more errors"
            
            await processing_msg.edit_text(result_msg, parse_mode='HTML')
            
            # Clear admin action
            context.user_data.clear()
            
            log_access(db, user, f"csv_upload_{success_count}_success_{error_count}_errors")
            
        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_path)
            except OSError:
                pass
    
    finally:
        db.close()


async def admin_upload_csv_callback(query, context, db, db_user):
    """Handle CSV upload request from admin."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_upload_csv")
    
    context.user_data['admin_action'] = 'upload_csv'
    
    message = (
        "📤 <b>Upload CSV File</b>\n\n"
        "Please upload a CSV file with the following columns:\n"
        "• <b>Range</b> - Name of the SMS range\n"
        "• <b>Number</b> - Phone number\n\n"
        "<i>Example CSV format:</i>\n"
        "<code>Range,Number\n"
        "Russia Lion Whatsapp,79032454671\n"
        "Russia Lion Whatsapp,79393992881</code>\n\n"
        "Other columns will be ignored.\n\n"
        "📎 Send the CSV file now:"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_manage_ranges_callback(query, context, db, db_user):
    """Show all ranges for admin management."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    log_access(db, db_user, "admin_manage_ranges")
    
    # Get all ranges
    all_ranges = get_all_ranges(db)
    
    message = "📋 <b>Manage Ranges & Prices</b>\n\n"
    
    if not all_ranges:
        message += "❌ No ranges uploaded yet.\n\n"
        message += "Please upload a CSV file first."
    else:
        message += f"Total Ranges: {len(all_ranges)}\n\n"
        message += "Click on a range to set price or delete:"
    
    # Create keyboard with ranges
    keyboard = []
    
    for range_obj, number_count in all_ranges[:15]:  # Show first 15
        # Get current price
        price = get_price_for_range(db, range_obj.unique_id)
        button_text = f"{range_obj.name} (${price:.2f}) - {number_count} nums"
        
        if len(button_text) > MAX_BUTTON_TEXT_LENGTH:
            button_text = button_text[:MAX_BUTTON_TEXT_LENGTH-3] + "..."
        
        keyboard.append([InlineKeyboardButton(
            button_text,
            callback_data=f"range_{range_obj.unique_id}"
        )])
    
    if len(all_ranges) > 15:
        keyboard.append([InlineKeyboardButton(
            f"... and {len(all_ranges) - 15} more ranges",
            callback_data="view_sms_ranges"
        )])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back to Admin Panel", callback_data="admin_back")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_set_range_price_callback(query, context, db, db_user, range_unique_id):
    """Prompt admin to set price for a range."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    range_obj = get_range_by_unique_id(db, range_unique_id)
    if not range_obj:
        await query.answer("❌ Range not found")
        return
    
    context.user_data['admin_action'] = 'set_range_price'
    context.user_data['range_unique_id'] = range_unique_id
    context.user_data['range_name'] = range_obj.name
    
    current_price = get_price_for_range(db, range_unique_id)
    
    message = (
        f"💵 <b>Set Price for Range</b>\n\n"
        f"<b>Range:</b> {escape_html(range_obj.name)}\n"
        f"<b>Current Price:</b> ${current_price:.2f}\n\n"
        f"Please send the new price (e.g., 1.50):"
    )
    
    keyboard = [[InlineKeyboardButton("❌ Cancel", callback_data=f"range_{range_unique_id}")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def admin_delete_range_callback(query, context, db, db_user, range_unique_id):
    """Delete a range and all its numbers."""
    if not is_user_admin(db, db_user.telegram_id):
        await query.answer("❌ Admin access required", show_alert=True)
        return
    
    range_obj = get_range_by_unique_id(db, range_unique_id)
    if not range_obj:
        await query.answer("❌ Range not found")
        return
    
    # Delete the range
    success = delete_range_and_numbers(db, range_unique_id)
    
    if success:
        message = (
            f"✅ <b>Range Deleted</b>\n\n"
            f"Range '<b>{escape_html(range_obj.name)}</b>' and all its numbers have been deleted."
        )
    else:
        message = "❌ Failed to delete range."
    
    keyboard = [[InlineKeyboardButton("⬅️ Back to Ranges", callback_data="admin_manage_ranges")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(message, reply_markup=reply_markup, parse_mode='HTML')


async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle errors in a clean and informative way."""
    # Extract useful information from update
    update_info = "No update information"
    if update:
        try:
            if update.effective_user:
                user_id = update.effective_user.id
                username = update.effective_user.username or "N/A"
                update_info = f"User {user_id} (@{username})"
            
            if update.effective_chat:
                chat_id = update.effective_chat.id
                update_info += f", Chat {chat_id}"
            
            if update.message and update.message.text:
                text = update.message.text[:100]  # Truncate long messages
                update_info += f", Text: '{text}'"
            elif update.callback_query:
                data = update.callback_query.data
                update_info += f", Callback: '{data}'"
        except Exception as e:
            update_info += f" (Error extracting info: {e})"
    
    # Log the error with clean information
    error_msg = f"Error occurred: {context.error.__class__.__name__}: {str(context.error)}"
    logger.error(f"{update_info} - {error_msg}")
    
    # Log full traceback at debug level
    if context.error:
        tb_list = traceback.format_exception(None, context.error, context.error.__traceback__)
        tb_string = ''.join(tb_list)
        logger.debug(f"Full traceback:\n{tb_string}")


def main():
    """Start the bot."""
    # Create the Application
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Register command handlers
    application.add_handler(CommandHandler("start", start_command))
    application.add_handler(CommandHandler("admin", admin_command))
    
    # Register callback query handler
    application.add_handler(CallbackQueryHandler(button_callback))
    
    # Register document handler for CSV uploads
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # Register message handler for text messages (admin actions and phone search)
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Register error handler
    application.add_error_handler(error_handler)
    
    # Start the bot
    logger.info("Starting ARGSMS Telegram Bot...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
